import shutil
import unittest
from unittest.mock import MagicMock, patch
import os
from openpyxl import Workbook
import pandas as pd
from utils.xlsx_chunk_reader import read_xlsx_in_chunks

class TestReadXlsxInChunks(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.test_data_dir="test_data"
        os.mkdir(cls.test_data_dir)
        # Create multiple temporary Excel files for testing
        cls.test_files = [f"{cls.test_data_dir}/test1.xlsx", f"{cls.test_data_dir}/test2.xlsx"]
        for file in cls.test_files:
            wb = Workbook()
            ws = wb.active
            ws.append(["A", "B", "C", "D"])  # Header
            for i in range(1, 5001):  # Add 5000 rows of data
                ws.append([i, i*2, i*3, i*4])
            wb.save(file)

    @classmethod
    def tearDownClass(cls):
        # Remove test files after tests
        shutil.rmtree(cls.test_data_dir)
        # os.remove(cls.test_data_dir)
        # for file in cls.test_files:
        #     if os.path.exists(file):
        #         os.remove(file)

    def test_chunk_readers(self):
        chunk_size = 1000
        chunk_readers = [read_xlsx_in_chunks(path, chunk_size=chunk_size) for path in self.test_files]
        
        for reader in chunk_readers:
            chunk_count = 0
            for chunk in reader:
                chunk_count += 1
                self.assertEqual(len(chunk), chunk_size if chunk_count < 5 else 1000)
                if chunk_count >= 5:
                    break  # Avoid infinite loops
            self.assertEqual(chunk_count, 5)  # Ensure all chunks are processed

    def test_chunk_combine(self):
        chunk_size = 1000
        chunk_readers = [read_xlsx_in_chunks(path, chunk_size=chunk_size) for path in self.test_files]
        
        df = pd.DataFrame()
        for reader in chunk_readers:
            chunk_count = 0
            for chunk in reader:
                chunk_count += 1

                df = pd.concat([df, chunk])
                # self.assertEqual(len(chunk), chunk_size if chunk_count < 5 else 1000)

                if chunk_count >= 5:
                    break  # Avoid infinite loops

        self.assertEqual(len(df.index), 10000) #Both test xlsx files are 5000 rows so the combined DataFrame's row count should be 1000
        
    def test_column_names(self):

        chunk = next(read_xlsx_in_chunks(self.test_files[0], chunk_size=1000))
        self.assertListEqual(list(chunk.columns), ["A", "B", "C", "D"])

        chunk2 = next(read_xlsx_in_chunks(self.test_files[1], chunk_size=1000))
        self.assertListEqual(list(chunk2.columns), ["A", "B", "C", "D"])

    def test_usecols(self):
        chunk = next(read_xlsx_in_chunks(self.test_files[0], chunk_size=1000, usecols=["A", "C"]))
        self.assertListEqual(list(chunk.columns), ["A", "C"])

        chunk2 = next(read_xlsx_in_chunks(self.test_files[1], chunk_size=1000, usecols=["A", "C"]))
        self.assertListEqual(list(chunk2.columns), ["A", "C"])

    @patch("logging.getLogger")
    def test_empty_file(self, mock_logger):
        logger_instance = MagicMock()
        mock_logger.return_value = logger_instance
        empty_file = f"{self.test_data_dir}/empty.xlsx"
        wb = Workbook()
        wb.save(empty_file)
        chunks = list(read_xlsx_in_chunks(empty_file, chunk_size=1000))
        self.assertEqual(len(chunks), 0)
        logger_instance.warning.assert_called_with(f"No rows found in first chunk of '{empty_file}'.")
        os.remove(empty_file)

    @patch("logging.getLogger")
    def test_shorter_file(self, mock_logger):
        logger_instance = MagicMock()
        mock_logger.return_value = logger_instance
        short_file = f"{self.test_data_dir}/short.xlsx"
        wb = Workbook()
        ws = wb.active
        for i in range(1, 4501):  # Add 4500 rows of data
            ws.append([i, i*2, i*3, i*4])


        chunk_size = 1000
        chunk_readers = [read_xlsx_in_chunks(path, chunk_size=chunk_size) for path in self.test_files]

        # TODO NOT WORKING

        for reader in chunk_readers:
            chunk_count = 0
            for chunk in reader:
                chunk_count += 1
                if chunk_count > 5:
                    logger_instance.debug.assert_called_with(f"Reached end of file '{short_file}', no more rows.")
                    break  # Avoid infinite loops



if __name__ == "__main__":
    unittest.main()